import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
HTML_FILE = ROOT / "publications.html"
DATA_FILE = ROOT / "data" / "publications.xlsx"
BACKUP_FILE = ROOT / "publications.backup.html"

# Load workbook
wb = load_workbook(DATA_FILE, data_only=True)
ws = wb["publications"]
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

def val(row, name):
    col = headers.get(name)
    if not col:
        return ""
    v = ws.cell(row, col).value
    return "" if v is None else str(v).strip()

# Build DOI -> slug map
doi_to_slug = {}
for row in range(2, ws.max_row + 1):
    doi = val(row, "doi")
    slug = val(row, "slug")
    if doi and slug:
        doi_to_slug[doi.lower()] = slug

html = HTML_FILE.read_text(encoding="utf-8")

# Save backup before editing
BACKUP_FILE.write_text(html, encoding="utf-8")

# Find each publication entry
pattern = re.compile(
    r'(<li class="pub-entry".*?>.*?<a class="pub-cite" href="(?P<href>[^"]+)".*?>.*?</a>)(?P<links>\s*<div class="pub-links">.*?</div>)',
    re.DOTALL
)

updated = 0

def add_read_article(match):
    global updated
    whole_start = match.group(1)
    href = match.group("href").strip()
    links = match.group("links")

    doi = ""
    m = re.search(r'https?://doi\.org/(.+)', href, re.I)
    if m:
        doi = m.group(1).strip().lower()

    if not doi:
        return match.group(0)

    slug = doi_to_slug.get(doi)
    if not slug:
        return match.group(0)

    # Avoid duplicate insertion
    if f'publications/{slug}.html' in links:
        return match.group(0)

    read_article_button = f'\n                  <a class="btn btn-ghost" href="publications/{slug}.html">Read article</a>'

    # Insert at the start of pub-links
    new_links = links.replace(
        '<div class="pub-links">',
        '<div class="pub-links">' + read_article_button,
        1
    )

    updated += 1
    return whole_start + new_links

new_html = pattern.sub(add_read_article, html)

HTML_FILE.write_text(new_html, encoding="utf-8")

print(f"Updated {updated} publication entries.")
print(f"Backup saved as: {BACKUP_FILE.name}")