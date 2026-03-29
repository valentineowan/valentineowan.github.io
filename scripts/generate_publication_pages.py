import html
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "publications.xlsx"
OUTPUT_DIR = ROOT / "publications"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

wb = load_workbook(DATA_FILE, data_only=True)
ws = wb.active

headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def val(row, name):
    col = headers.get(name)
    if not col:
        return ""
    v = ws.cell(row, col).value
    return "" if v is None else str(v).strip()


def esc(text):
    return html.escape(text or "")


def citation_js(text):
    if not text:
        return ""
    return text.replace("\\", "\\\\").replace("`", "\\`")


def split_authors_for_meta(authors_text):
    if not authors_text:
        return []
    parts = [a.strip() for a in authors_text.split(";") if a.strip()]
    return parts


def build_meta_tags(title, authors, year, journal, source_title, doi, pdf_path, slug):
    meta = []

    if title:
        meta.append(f'<meta name="citation_title" content="{esc(title)}" />')

    for author in split_authors_for_meta(authors):
        meta.append(f'<meta name="citation_author" content="{esc(author)}" />')

    if year:
        meta.append(f'<meta name="citation_publication_date" content="{esc(year)}" />')

    journal_title = journal or source_title
    if journal_title:
        meta.append(f'<meta name="citation_journal_title" content="{esc(journal_title)}" />')

    if doi:
        meta.append(f'<meta name="citation_doi" content="{esc(doi)}" />')

    if pdf_path:
        meta.append(f'<meta name="citation_pdf_url" content="../{esc(pdf_path)}" />')

    meta.append(
        f'<link rel="canonical" href="https://www.valentineowan.com/publications/{esc(slug)}.html" />'
    )

    return "\n  ".join(meta)


def render_page(row):
    slug = val(row, "slug")
    if not slug:
        return None, None

    title = val(row, "title")
    authors = val(row, "authors")
    year = val(row, "year")
    source_title = val(row, "source_title")
    journal = val(row, "journal")
    booktitle = val(row, "booktitle")
    publisher = val(row, "publisher")
    school = val(row, "school")
    volume = val(row, "volume")
    issue = val(row, "issue")
    pages = val(row, "pages")
    doi = val(row, "doi")
    abstract = val(row, "abstract")
    keywords = val(row, "keywords")
    pdf_path = val(row, "pdf_path")
    access = val(row, "access").lower()
    entry_type = val(row, "entry_type")

    apa = val(row, "apa_citation")
    mla = val(row, "mla_citation")
    harvard = val(row, "harvard_citation")
    vancouver = val(row, "vancouver_citation")
    bibtex = val(row, "bibtex")

    openalex_citations = val(row, "openalex_citations")
    openalex_id = val(row, "openalex_id")
    openalex_url = val(row, "openalex_url")
    openalex_last_checked = val(row, "openalex_last_checked")

    source_parts = []
    if journal:
        source_parts.append(f"<p><strong>Journal:</strong> {esc(journal)}</p>")
    if booktitle:
        source_parts.append(f"<p><strong>Book title:</strong> {esc(booktitle)}</p>")
    if publisher:
        source_parts.append(f"<p><strong>Publisher:</strong> {esc(publisher)}</p>")
    if school:
        source_parts.append(f"<p><strong>Institution:</strong> {esc(school)}</p>")
    if year:
        source_parts.append(f"<p><strong>Year:</strong> {esc(year)}</p>")
    if volume:
        source_parts.append(f"<p><strong>Volume:</strong> {esc(volume)}</p>")
    if issue:
        source_parts.append(f"<p><strong>Issue:</strong> {esc(issue)}</p>")
    if pages:
        source_parts.append(f"<p><strong>Pages / article number:</strong> {esc(pages)}</p>")
    if entry_type:
        source_parts.append(f"<p><strong>Type:</strong> {esc(entry_type)}</p>")
    if doi:
        source_parts.append(
            f'<p><strong>DOI:</strong> <a href="https://doi.org/{esc(doi)}" target="_blank" rel="noopener">{esc(doi)}</a></p>'
        )

    actions = []
    if pdf_path and access == "open":
        actions.append(
            f'<a class="btn" href="../{esc(pdf_path)}" target="_blank" rel="noopener">Read PDF</a>'
        )
    if doi:
        actions.append(
            f'<a class="btn btn-ghost" href="https://doi.org/{esc(doi)}" target="_blank" rel="noopener">View DOI</a>'
        )
    if access == "request":
        actions.append(
            '<a class="btn btn-ghost" href="../contact.html">Request a copy</a>'
        )
    actions.append(
        '<a class="btn btn-ghost" href="../publications.html">Back to publications</a>'
    )

    keyword_block = ""
    if keywords:
        keyword_block = f"""
        <section class="card article-block">
          <h2>Keywords</h2>
          <p>{esc(keywords)}</p>
        </section>
        """

    abstract_block = ""
    if abstract:
        abstract_block = f"""
        <section class="card article-block">
          <h2>Abstract</h2>
          <p>{esc(abstract)}</p>
        </section>
        """

    citation_count_block = ""
    if openalex_citations and openalex_citations not in {"None", ""}:
        openalex_link = ""
        if openalex_url:
            openalex_link = f' <a href="{esc(openalex_url)}" target="_blank" rel="noopener">(view record)</a>'

        checked_line = ""
        if openalex_last_checked:
            checked_line = f'<p><small>Last checked: {esc(openalex_last_checked)}</small></p>'

        citation_count_block = f"""
        <section class="card article-block">
          <h2>Citation metrics</h2>
          <p><strong>Citations:</strong> {esc(openalex_citations)}{openalex_link}</p>
          <p><small>Source: OpenAlex</small></p>
          {checked_line}
        </section>
        """

    source_block = "\n".join(source_parts)
    actions_block = "\n".join(actions)
    meta_tags = build_meta_tags(
        title=title,
        authors=authors,
        year=year,
        journal=journal,
        source_title=source_title,
        doi=doi,
        pdf_path=pdf_path,
        slug=slug,
    )

    page = f"""<!doctype html>
<html lang="en-GB">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{esc(title)} | Valentine Joseph Owan</title>
  <meta name="description" content="{esc(title)}" />
  {meta_tags}
  <link rel="stylesheet" href="../assets/css/style.css" />
  <style>
    .article-wrap {{
      display:grid;
      gap:16px;
    }}
    .article-top h1 {{
      margin-bottom:10px;
    }}
    .article-authors {{
      color:var(--muted);
      font-size:1rem;
    }}
    .article-actions {{
      display:flex;
      flex-wrap:wrap;
      gap:10px;
      margin-top:14px;
    }}
    .article-meta p {{
      margin:0 0 8px;
      color:var(--muted);
    }}
    .article-block h2 {{
      margin-top:0;
    }}
    .citation-tabs {{
      display:flex;
      flex-wrap:wrap;
      gap:8px;
      margin:10px 0 12px;
    }}
    .citation-tabs button {{
      padding:8px 12px;
      border-radius:10px;
      border:1px solid var(--line);
      background:var(--panel);
      cursor:pointer;
      color:var(--text);
    }}
    .citation-tabs button:hover {{
      background:rgba(17,24,39,.05);
    }}
    .citation-box {{
      white-space:pre-wrap;
      overflow-wrap:anywhere;
      color:var(--text);
    }}
    .crumbs {{
      font-size:14px;
      color:var(--muted);
      margin-bottom:12px;
    }}
    .crumbs a {{
      text-decoration:none;
    }}
  </style>
</head>
<body>
  <a class="skip-link" href="#main">Skip to content</a>

  <header class="site-header">
    <div class="container header-inner">
      <div class="brand">
        <a href="../index.html" class="brand-name">Valentine Joseph Owan</a>
        <div class="brand-tag">Researcher | Psychometrician | Statistician</div>
      </div>

      <button class="nav-toggle" type="button" aria-expanded="false" aria-controls="site-nav">
        Menu
      </button>

      <nav id="site-nav" class="site-nav" aria-label="Primary">
        <a class="nav-link" href="../index.html">Home</a>
        <a class="nav-link" href="../about.html">About</a>
        <a class="nav-link active" href="../publications.html">Publications</a>
        <a class="nav-link" href="../teaching.html">Teaching</a>
        <a class="nav-link" href="../cv.html">CV</a>
        <a class="nav-link" href="../updates.html">Updates</a>
        <a class="nav-link" href="../contact.html">Contact</a>
        <a class="nav-link" href="../quotes.html">Quotes</a>
      </nav>
    </div>
  </header>

  <main id="main" class="section">
    <div class="container narrow">
      <div class="crumbs">
        <a href="../index.html">Home</a> / <a href="../publications.html">Publications</a> / Record
      </div>

      <div class="article-wrap">
        <section class="card article-top">
          <p class="kicker">Publication record</p>
          <h1>{esc(title)}</h1>
          <p class="article-authors">{esc(authors)}</p>
          <div class="article-actions">
            {actions_block}
          </div>
        </section>

        <section class="card article-meta">
          <h2>Publication details</h2>
          {source_block}
        </section>

        {citation_count_block}

        {abstract_block}

        {keyword_block}

        <section class="card article-block">
          <h2>Citation</h2>

          <div class="citation-tabs">
            <button onclick="showCitation('apa')">APA</button>
            <button onclick="showCitation('mla')">MLA</button>
            <button onclick="showCitation('harvard')">Harvard</button>
            <button onclick="showCitation('vancouver')">Vancouver</button>
            <button onclick="showCitation('bibtex')">BibTeX</button>
          </div>

          <div id="citationText" class="citation-box"></div>

          <div class="card-actions">
            <button class="btn btn-ghost" onclick="copyCitation()">Copy citation</button>
          </div>
        </section>
      </div>
    </div>
  </main>

  <footer class="site-footer">
    <div class="container">
      <div class="footer-top">
        <div class="footer-copy">
          © <span id="year"></span> Valentine Joseph Owan
          <span class="footer-sep">•</span>
          University of Calabar
        </div>

        <nav class="footer-links" aria-label="Footer">
          <a href="../index.html">Home</a>
          <a href="../about.html">About</a>
          <a href="../publications.html">Publications</a>
          <a href="../teaching.html">Teaching</a>
          <a href="../cv.html">CV</a>
          <a href="../updates.html">Updates</a>
          <a href="../quotes.html">Quotes</a>
          <a href="../contact.html">Contact</a>
        </nav>
      </div>

      <div class="footer-note">
        Academic website for research, teaching, and scholarly work.
      </div>
    </div>
  </footer>

  <script src="../assets/js/main.js"></script>
  <script>
    const citations = {{
      apa: `{citation_js(apa)}`,
      mla: `{citation_js(mla)}`,
      harvard: `{citation_js(harvard)}`,
      vancouver: `{citation_js(vancouver)}`,
      bibtex: `{citation_js(bibtex)}`
    }};

    function showCitation(type) {{
      const box = document.getElementById("citationText");
      if (!box) return;
      box.innerHTML = citations[type] || "";
    }}

    function copyCitation() {{
      const box = document.getElementById("citationText");
      if (!box) return;
      const text = box.innerText;
      navigator.clipboard.writeText(text).then(() => {{
        alert("Citation copied");
      }});
    }}

    (function () {{
      var y = document.getElementById("year");
      if (y) y.textContent = new Date().getFullYear();
      showCitation("apa");
    }})();
  </script>
</body>
</html>
"""
    return slug, page


generated = 0

for row in range(2, ws.max_row + 1):
    slug, page = render_page(row)
    if not slug or not page:
        continue
    out_file = OUTPUT_DIR / f"{slug}.html"
    out_file.write_text(page, encoding="utf-8")
    generated += 1

print(f"Generated {generated} publication pages in {OUTPUT_DIR}")